"""
    Split recording into windows and get mfcc for each window, as well as derivatives. Pass each window throught an svm
        and vote to see overall what the accent is.
"""

import openpyxl


class svm_accent_classifier:
    def __init__(self, dataset_path='VietMed'):
        self.metadata_workbook_path = dataset_path + '/Metadata_labeled_medical.xlsx'

    def preprocessor(self):
        self.metadata_workbook = openpyxl.load_workbook(self.metadata_workbook_path)
        sheet_obj = self.metadata_workbook.active
        max_row = sheet_obj.max_row 
        max_col = sheet_obj.max_column 

        print(f'{max_row}, {max_col}')
        
        # Loop over the rows and create a dictionary of recordings. Not sure how to store people separately.
        self.metadata = []
        for i in range(1, max_row):
            cell_obj = sheet_obj.cell(row=i, column=3)
            if cell_obj.value != None:
                self.metadata.append({'Name': cell_obj.value, })

        # Create a dictionary of the file names of each recording and all of the metadata that goes with it.
        # Split it up into training and test data.
        # Because each Audio name contains multiple recordings with multiple people, make sure that the recordings are labelled correctly


if __name__ == '__main__':
    classifier = svm_accent_classifier()
    classifier.preprocessor()